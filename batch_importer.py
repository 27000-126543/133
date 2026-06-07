import os
import re
import uuid
from datetime import datetime, date
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass, field
import pandas as pd

from database import get_db_session
from models import Employee, Certificate, CertificateType, Department, ImportBatch
from config import Config


@dataclass
class ImportRecord:
    row_index: int
    data: Dict[str, Any]
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    is_valid: bool = True
    is_duplicate: bool = False
    certificate_id: Optional[int] = None
    employee_id: Optional[int] = None


@dataclass
class ImportResult:
    batch_no: str
    total_records: int
    success_count: int
    duplicate_count: int
    error_count: int
    error_details: List[Dict]
    success_records: List[ImportRecord]
    error_records: List[ImportRecord]


class BatchImporter:
    REQUIRED_COLUMNS = [
        "employee_no",
        "cert_name",
        "cert_number",
        "issuing_authority",
    ]
    
    OPTIONAL_COLUMNS = [
        "employee_name",
        "department",
        "position",
        "cert_type",
        "issue_date",
        "expiry_date",
        "score",
        "source",
    ]
    
    DATE_FORMATS = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y年%m月%d日",
        "%Y-%m-%d",
        "%Y.%m.%d",
    ]
    
    def __init__(self):
        self.batch_size = Config.BATCH_SIZE
    
    def _parse_date(self, date_str: str) -> Optional[date]:
        if not date_str or pd.isna(date_str):
            return None
        
        date_str = str(date_str).strip()
        
        for fmt in self.DATE_FORMATS:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
    
    def _generate_batch_no(self) -> str:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        unique = uuid.uuid4().hex[:6].upper()
        return f"IMP{timestamp}{unique}"
    
    def _validate_employee(self, record: ImportRecord, employee_map: Dict[str, Employee]) -> Optional[Employee]:
        data = record.data
        employee_no = str(data.get("employee_no", "")).strip()
        
        if not employee_no:
            record.errors.append("员工工号不能为空")
            record.is_valid = False
            return None
        
        employee = employee_map.get(employee_no)
        
        if not employee:
            employee_name = str(data.get("employee_name", "")).strip()
            department_name = str(data.get("department", "")).strip()
            position = str(data.get("position", "")).strip()
            
            if not employee_name:
                record.errors.append(f"员工工号{employee_no}不存在，且未提供员工姓名无法创建")
                record.is_valid = False
                return None
            
            with get_db_session() as session:
                dept = None
                if department_name:
                    dept = session.query(Department).filter(Department.name == department_name).first()
                    if not dept:
                        dept = Department(name=department_name)
                        session.add(dept)
                        session.flush()
                
                employee = Employee(
                    employee_no=employee_no,
                    name=employee_name,
                    department_id=dept.id if dept else None,
                    position=position if position else None,
                    is_active=True,
                )
                
                session.add(employee)
                session.flush()
                employee_id = employee.id
            
            with get_db_session() as session:
                employee = session.query(Employee).filter(Employee.id == employee_id).first()
                employee_map[employee_no] = employee
                record.warnings.append(f"员工{employee_no}不存在，已自动创建")
        else:
            record.employee_id = employee.id
        
        return employee
    
    def _validate_cert_type(self, record: ImportRecord, cert_type_map: Dict[str, CertificateType]) -> Optional[CertificateType]:
        data = record.data
        cert_name = str(data.get("cert_name", "")).strip()
        cert_type_name = str(data.get("cert_type", "")).strip()
        
        search_name = cert_type_name or cert_name
        
        if not search_name:
            record.errors.append("证书名称或类型不能为空")
            record.is_valid = False
            return None
        
        cert_type = cert_type_map.get(search_name)
        
        if not cert_type:
            for name, ct in cert_type_map.items():
                if search_name in name or name in search_name:
                    cert_type = ct
                    break
        
        if not cert_type:
            with get_db_session() as session:
                cert_type = CertificateType(
                    code=re.sub(r'[^A-Z0-9]', '', search_name.upper())[:20] or uuid.uuid4().hex[:10].upper(),
                    name=search_name,
                    category="其他",
                    is_required=False,
                )
                session.add(cert_type)
                session.flush()
                cert_type_id = cert_type.id
            
            with get_db_session() as session:
                cert_type = session.query(CertificateType).filter(CertificateType.id == cert_type_id).first()
                cert_type_map[search_name] = cert_type
                record.warnings.append(f"证书类型{search_name}不存在，已自动创建")
        
        return cert_type
    
    def _check_duplicate(self, employee_id: int, cert_type_id: int, cert_number: str, session) -> bool:
        if not cert_number:
            return False
        
        existing = session.query(Certificate).filter(
            Certificate.employee_id == employee_id,
            Certificate.cert_type_id == cert_type_id,
            Certificate.cert_number == cert_number,
            Certificate.status.in_(["valid", "expired"])
        ).first()
        
        return existing is not None
    
    def _validate_record(self, record: ImportRecord, employee_map: Dict[str, Employee], cert_type_map: Dict[str, CertificateType]) -> bool:
        data = record.data
        
        employee = self._validate_employee(record, employee_map)
        if not employee:
            return False
        
        cert_type = self._validate_cert_type(record, cert_type_map)
        if not cert_type:
            return False
        
        cert_number = str(data.get("cert_number", "")).strip()
        if not cert_number:
            record.errors.append("证书编号不能为空")
            record.is_valid = False
            return False
        
        issuing_authority = str(data.get("issuing_authority", "")).strip()
        if not issuing_authority:
            record.errors.append("颁发机构不能为空")
            record.is_valid = False
            return False
        
        issue_date = self._parse_date(data.get("issue_date"))
        expiry_date = self._parse_date(data.get("expiry_date"))
        
        if issue_date and expiry_date and issue_date >= expiry_date:
            record.errors.append("签发日期必须早于有效期截止日期")
            record.is_valid = False
            return False
        
        with get_db_session() as session:
            if self._check_duplicate(employee.id, cert_type.id, cert_number, session):
                record.is_duplicate = True
                record.is_valid = False
                record.warnings.append(f"证书{cert_number}已存在，跳过")
                return False
        
        score = data.get("score")
        if score is not None and not pd.isna(score):
            try:
                score = float(score)
                if score < 0 or score > 100:
                    record.errors.append("成绩应在0-100之间")
                    record.is_valid = False
                    return False
            except (ValueError, TypeError):
                record.errors.append("成绩格式不正确")
                record.is_valid = False
                return False
        
        record.data["_parsed_issue_date"] = issue_date
        record.data["_parsed_expiry_date"] = expiry_date
        record.data["_employee"] = employee
        record.data["_cert_type"] = cert_type
        
        return True
    
    def _import_record(self, record: ImportRecord) -> bool:
        data = record.data
        employee = data["_employee"]
        cert_type = data["_cert_type"]
        
        try:
            with get_db_session() as session:
                old_cert = session.query(Certificate).filter(
                    Certificate.employee_id == employee.id,
                    Certificate.cert_type_id == cert_type.id,
                    Certificate.status == "valid"
                ).first()
                
                cert = Certificate(
                    employee_id=employee.id,
                    cert_type_id=cert_type.id,
                    cert_name=str(data.get("cert_name", cert_type.name)).strip(),
                    cert_number=str(data.get("cert_number", "")).strip(),
                    issuing_authority=str(data.get("issuing_authority", "")).strip(),
                    issue_date=data.get("_parsed_issue_date"),
                    expiry_date=data.get("_parsed_expiry_date"),
                    score=float(data.get("score")) if data.get("score") is not None and not pd.isna(data.get("score")) else None,
                    source=str(data.get("source", "batch_import")).strip(),
                    status="valid",
                    verified=True,
                    verified_at=datetime.now(),
                )
                
                session.add(cert)
                session.flush()
                
                if old_cert and cert.expiry_date and old_cert.expiry_date:
                    if cert.expiry_date > old_cert.expiry_date:
                        old_cert.status = "replaced"
                
                record.certificate_id = cert.id
                record.employee_id = employee.id
                return True
                
        except Exception as e:
            record.errors.append(f"数据库保存失败: {str(e)}")
            return False
    
    def import_from_file(self, file_path: str, imported_by: Optional[int] = None) -> ImportResult:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == ".xlsx" or file_ext == ".xls":
            df = pd.read_excel(file_path, dtype=str)
        elif file_ext == ".csv":
            df = pd.read_csv(file_path, dtype=str, encoding="utf-8-sig")
        else:
            raise ValueError(f"不支持的文件格式: {file_ext}，仅支持 .xlsx, .xls, .csv")
        
        df.columns = [col.lower().strip().replace(" ", "_") for col in df.columns]
        
        missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
        if missing_columns:
            raise ValueError(f"缺少必需列: {', '.join(missing_columns)}")
        
        batch_no = self._generate_batch_no()
        
        with get_db_session() as session:
            employees = session.query(Employee).all()
            employee_map = {e.employee_no: e for e in employees}
            
            cert_types = session.query(CertificateType).all()
            cert_type_map = {ct.name: ct for ct in cert_types}
        
        records = []
        for idx, row in df.iterrows():
            record = ImportRecord(
                row_index=idx + 2,
                data=row.to_dict(),
            )
            records.append(record)
        
        for record in records:
            self._validate_record(record, employee_map, cert_type_map)
        
        valid_records = [r for r in records if r.is_valid and not r.is_duplicate]
        duplicate_records = [r for r in records if r.is_duplicate]
        error_records = [r for r in records if not r.is_valid and not r.is_duplicate]
        
        success_count = 0
        for i in range(0, len(valid_records), self.batch_size):
            batch = valid_records[i:i + self.batch_size]
            for record in batch:
                if self._import_record(record):
                    success_count += 1
        
        success_records = [r for r in valid_records if r.certificate_id]
        error_records.extend([r for r in valid_records if not r.certificate_id])
        
        error_details = []
        for record in error_records:
            error_details.append({
                "row": record.row_index,
                "employee_no": record.data.get("employee_no"),
                "cert_name": record.data.get("cert_name"),
                "cert_number": record.data.get("cert_number"),
                "errors": record.errors,
                "warnings": record.warnings,
            })
        
        with get_db_session() as session:
            import_batch = ImportBatch(
                batch_no=batch_no,
                file_name=os.path.basename(file_path),
                total_records=len(records),
                success_count=success_count,
                duplicate_count=len(duplicate_records),
                error_count=len(error_records),
                error_details=error_details,
                status="completed",
                imported_by=imported_by,
                completed_at=datetime.now(),
            )
            session.add(import_batch)
        
        return ImportResult(
            batch_no=batch_no,
            total_records=len(records),
            success_count=success_count,
            duplicate_count=len(duplicate_records),
            error_count=len(error_records),
            error_details=error_details,
            success_records=success_records,
            error_records=error_records,
        )
    
    def generate_import_template(self, output_path: str) -> str:
        columns = self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS
        
        column_aliases = {
            "employee_no": "员工工号",
            "employee_name": "员工姓名",
            "department": "部门",
            "position": "岗位",
            "cert_name": "证书名称",
            "cert_type": "证书类型",
            "cert_number": "证书编号",
            "issuing_authority": "颁发机构",
            "issue_date": "签发日期",
            "expiry_date": "有效期至",
            "score": "成绩",
            "source": "来源",
        }
        
        display_columns = [column_aliases.get(col, col) for col in columns]
        
        sample_data = [
            ["E001", "张三", "技术部", "项目经理", "一级建造师", "注册建造师", "JZS2023001234",
             "住房和城乡建设部", "2023-03-15", "2026-03-14", "", "manual"],
            ["E002", "李四", "安全部", "安全员", "安全员B证", "安全生产考核", "京建安B2023000567",
             "北京市住房和城乡建设委员会", "2023-06-20", "2026-06-19", "", "manual"],
        ]
        
        df = pd.DataFrame(sample_data, columns=display_columns)
        
        file_ext = os.path.splitext(output_path)[1].lower()
        if file_ext == ".xlsx":
            writer = pd.ExcelWriter(output_path, engine="openpyxl")
            df.to_excel(writer, index=False, sheet_name="证书导入模板")
            
            workbook = writer.book
            worksheet = writer.sheets["证书导入模板"]
            
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for col_idx, col_name in enumerate(display_columns, 1):
                if col_idx <= len(self.REQUIRED_COLUMNS):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = f"*{col_name}"
            
            writer.close()
        elif file_ext == ".csv":
            df.to_csv(output_path, index=False, encoding="utf-8-sig")
        else:
            raise ValueError(f"不支持的文件格式: {file_ext}")
        
        return output_path


_batch_importer_singleton = None


def get_batch_importer() -> BatchImporter:
    global _batch_importer_singleton
    if _batch_importer_singleton is None:
        _batch_importer_singleton = BatchImporter()
    return _batch_importer_singleton
